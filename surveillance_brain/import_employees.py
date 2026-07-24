#!/usr/bin/env python3
"""
Bulk-import employees from an Excel roster with face embeddings.
======================================================================
Columns (header names matched loosely; extra columns ignored):
    employee code | name | designation | photo

Photos can come from EITHER:
  • images EMBEDDED in the .xlsx (pasted into the sheet) — matched to each row by
    the drawing anchor (this is what farmlist_all1.xlsx uses); or
  • filenames listed in the `photo` cell (comma/;/newline separated), resolved in
    the same --folder.
Both sources are merged, so a row can have several photos either way.

For each row we extract face embeddings with the live SCRFD+AdaFace model (via the
AI venv) and enroll — best face = primary embedding, the rest = extra gallery views,
plus a durable locked profile photo. `employee code` is the idempotency key:
re-running UPDATES an existing employee instead of duplicating.

Run from this directory, in the Brain venv (Postgres + Qdrant must be up):

    cd surveillance_brain
    .venv/bin/python import_employees.py --xlsx /home/antz/Desktop/farmlist_all1.xlsx           # DRY RUN
    .venv/bin/python import_employees.py --xlsx /home/antz/Desktop/farmlist_all1.xlsx --apply    # import
"""

from __future__ import annotations

import argparse
import asyncio
import base64
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict
from xml.etree import ElementTree as ET

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))   # make brain modules importable

from openpyxl import load_workbook   # noqa: E402

# Loose header matching — exact alias first, then "header contains alias" (longest
# alias wins), so 'Payroll Company Employee Code' still maps to the code column.
ALIASES = {
    "code":  ["employee code", "employee_code", "emp code", "employee id", "employee_id",
              "external_id", "empid", "code", "id"],
    "name":  ["employee name", "full name", "emp name", "name"],
    "dept":  ["designation", "department", "dept", "role", "title"],
    "photo": ["photo(s)", "photos", "photo", "image", "images", "picture", "pictures"],
}

_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip().lower())


def map_columns(header) -> dict:
    norm = [_norm(h) for h in header]
    idx = {}
    for key, aliases in ALIASES.items():
        hit = next((norm.index(a) for a in aliases if a in norm), None)   # exact
        if hit is None:                                                   # else substring
            for a in sorted(aliases, key=len, reverse=True):
                hit = next((i for i, h in enumerate(norm) if a in h), None)
                if hit is not None:
                    break
        if hit is not None:
            idx[key] = hit
    return idx


def split_photos(cell) -> list:
    if cell is None:
        return []
    return [p.strip() for p in re.split(r"[,;\n]+", str(cell)) if p.strip()]


def extract_embedded_by_row(xlsx: str) -> dict:
    """Map spreadsheet row (0-indexed) -> [media names] for images embedded in the
    sheet, by reading the drawing anchors. Returns {} if there are no embedded
    images. Best-effort — never raises."""
    byrow = defaultdict(list)
    try:
        z = zipfile.ZipFile(xlsx)
    except Exception:
        return {}
    try:
        drawings = [n for n in z.namelist() if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n)]
        for dpath in drawings:
            relpath = os.path.join(os.path.dirname(dpath), "_rels", os.path.basename(dpath) + ".rels")
            rid2media = {}
            if relpath in z.namelist():
                rels = ET.fromstring(z.read(relpath))
                for rel in rels:
                    tgt = rel.get("Target", "")
                    rid2media[rel.get("Id")] = re.sub(r"^\.\./", "xl/", tgt)
            dr = ET.fromstring(z.read(dpath))
            for tag in ("twoCellAnchor", "oneCellAnchor"):
                for anc in dr.findall(f"xdr:{tag}", _NS):
                    frm = anc.find("xdr:from", _NS)
                    if frm is None:
                        continue
                    row = int(frm.find("xdr:row", _NS).text)
                    blip = anc.find(".//a:blip", _NS)
                    emb = blip.get("{%s}embed" % _NS["r"]) if blip is not None else None
                    media = rid2media.get(emb)
                    if media:
                        byrow[row].append(media)
    except Exception:
        return {}
    finally:
        z.close()
    return dict(byrow)


def read_rows(xlsx: str, folder: str) -> list:
    embedded_by_row = extract_embedded_by_row(xlsx)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        raise SystemExit("Empty spreadsheet.")
    idx = map_columns(header)
    missing = [k for k in ("code", "name") if k not in idx]
    if missing:
        raise SystemExit(f"Missing required column(s) {missing}. Found headers: {list(header)}")

    rows = []
    for di, r in enumerate(it):                 # di = 0-based DATA index
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        sheet_row = di + 1                       # 0-indexed spreadsheet row (header is row 0)
        code = str(r[idx["code"]]).strip() if r[idx["code"]] is not None else ""
        name = str(r[idx["name"]]).strip() if r[idx["name"]] is not None else ""
        dept = ""
        if "dept" in idx and r[idx["dept"]] is not None:
            dept = str(r[idx["dept"]]).strip()
        dept = dept or "General"

        embedded = list(embedded_by_row.get(sheet_row, []))     # media names inside the xlsx
        files, miss = [], []
        if "photo" in idx:                                      # optional filename column
            for pn in split_photos(r[idx["photo"]]):
                fp = os.path.join(folder, pn)
                if os.path.isfile(fp):
                    files.append(fp)
                else:
                    base = os.path.basename(pn).lower()
                    hit = next((os.path.join(folder, f) for f in os.listdir(folder)
                                if f.lower() == base), None) if os.path.isdir(folder) else None
                    (files.append(hit) if hit else miss.append(pn))

        n_photos = len(embedded) + len(files)
        errs = []
        if not code:
            errs.append("missing employee code")
        if not name:
            errs.append("missing name")
        if n_photos == 0:
            errs.append("no photo (none embedded on this row" +
                        (f"; not found: {', '.join(miss)}" if miss else "") + ")")
        rows.append({"code": code, "name": name, "dept": dept,
                     "embedded": embedded, "files": files, "missing": miss,
                     "n_photos": n_photos, "errors": errs})

    # Idempotency key = employee code, BUT some codes are placeholders reused across
    # people (e.g. 'Stipend' ×11). A shared key would merge those people into one on
    # import, so any DUPLICATE code becomes a per-person key 'code:name' instead.
    counts = Counter(r["code"] for r in rows if r["code"])
    for r in rows:
        if r["code"] and counts[r["code"]] > 1:
            r["external_id"] = f"{r['code']}:{r['name']}"
            r["synth"] = True
        else:
            r["external_id"] = r["code"]
            r["synth"] = False
    return rows


def _images_b64(xlsx: str, row: dict) -> list:
    """Build the base64 image list for a row from embedded media + external files."""
    out = []
    if row["embedded"]:
        with zipfile.ZipFile(xlsx) as z:
            for media in row["embedded"]:
                try:
                    out.append(base64.b64encode(z.read(media)).decode())
                except KeyError:
                    pass
    for fp in row["files"]:
        with open(fp, "rb") as f:
            out.append(base64.b64encode(f.read()).decode())
    return out


async def apply_rows(xlsx: str, rows: list) -> None:
    from services import enrollment_service
    created = updated = skipped = 0
    for row in rows:
        if row["errors"]:
            skipped += 1
            print(f"  SKIP   {row['code'] or '—':<12} {row['name'][:24]:<24} — {'; '.join(row['errors'])}")
            continue
        try:
            imgs = _images_b64(xlsx, row)
            res = await enrollment_service.import_employee_from_images(
                external_id=row["external_id"], name=row["name"], department=row["dept"], images_b64=imgs)
            created += res["action"] == "created"
            updated += res["action"] == "updated"
            print(f"  {res['action'].upper():<7}{row['code']:<12} {row['name'][:24]:<24} "
                  f"id={res['identity_id']}  {res['photos_used']}/{row['n_photos']} face(s)")
        except Exception as e:  # noqa: BLE001
            skipped += 1
            print(f"  FAIL   {row['code']:<12} {row['name'][:24]:<24} — {e}")
    print(f"\nDone: {created} created, {updated} updated, {skipped} skipped.")


def main() -> None:
    ap = argparse.ArgumentParser(description="Bulk-import employees (Excel + embedded/foldered photos) with face embeddings.")
    ap.add_argument("--xlsx", default=None, help="path to the .xlsx roster")
    ap.add_argument("--folder", default=None, help="folder holding the .xlsx (and any filename photos)")
    ap.add_argument("--only", default=None,
                    help="import ONLY rows whose employee code equals, or whose name contains, "
                         "this value (case-insensitive) — e.g. --only G338 or --only 'Carol Rebello'")
    ap.add_argument("--apply", action="store_true", help="actually import (default: dry-run preview)")
    args = ap.parse_args()

    if args.xlsx:
        xlsx = os.path.abspath(args.xlsx)
        folder = args.folder and os.path.abspath(args.folder) or os.path.dirname(xlsx)
    elif args.folder:
        folder = os.path.abspath(args.folder)
        xs = [f for f in os.listdir(folder) if f.lower().endswith((".xlsx", ".xlsm"))]
        if len(xs) != 1:
            raise SystemExit(f"expected exactly one .xlsx in {folder}, found {xs}. Pass --xlsx.")
        xlsx = os.path.join(folder, xs[0])
    else:
        raise SystemExit("give --xlsx PATH (or --folder containing one .xlsx)")
    if not os.path.isfile(xlsx):
        raise SystemExit(f"xlsx not found: {xlsx}")

    rows = read_rows(xlsx, folder)
    if args.only:
        q = args.only.strip().lower()
        rows = [r for r in rows if r["code"].lower() == q or q in r["name"].lower()]
        if not rows:
            raise SystemExit(f"--only {args.only!r} matched no rows in the roster.")
        print(f"--only {args.only!r}: {len(rows)} matching row(s).\n")
    ready = [r for r in rows if not r["errors"]]
    embedded_total = sum(len(r["embedded"]) for r in rows)
    print(f"Roster : {xlsx}")
    print(f"{len(rows)} row(s): {len(ready)} ready, {len(rows) - len(ready)} with errors "
          f"| {embedded_total} embedded photo(s)\n")
    for i, r in enumerate(rows, 1):
        status = "READY" if not r["errors"] else "ERROR: " + "; ".join(r["errors"])
        key = f"  key={r['external_id']}" if r.get("synth") else ""
        print(f"#{i:>3}  {r['code'] or '—':<12} {r['name'][:22]:<22} "
              f"dept={r['dept'][:18]:<18} photos={r['n_photos']}  {status}{key}")

    if not args.apply:
        print("\nDRY RUN — nothing written. Re-run with --apply to import.")
        return
    if not ready:
        print("\nNothing ready to import.")
        return
    asyncio.run(apply_rows(xlsx, rows))


if __name__ == "__main__":
    main()
